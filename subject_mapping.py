#!/usr/bin/env python3
"""
科目名称映射配置
从JSON配置文件读取科目映射规则，支持灵活匹配
"""

import json
import os

# 配置文件路径
CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'subject_mapping_config.json')

def load_config():
    """
    从JSON配置文件加载科目映射规则

    Returns:
        dict: 配置数据
    """
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"科目映射配置文件未找到: {CONFIG_FILE}")
    except json.JSONDecodeError as e:
        raise ValueError(f"科目映射配置文件格式错误: {e}")

def build_mapping_dict(subjects_config):
    """
    从JSON配置构建映射字典

    Args:
        subjects_config: 科目配置字典

    Returns:
        dict: 格式为 {标准名称: [别名列表]}
    """
    mapping = {}
    for standard_name, config in subjects_config.items():
        mapping[standard_name] = config['aliases']
    return mapping

# 加载配置
config = None
try:
    config = load_config()

    # 构建映射字典（保持与原格式兼容）
    BALANCE_MAPPING = build_mapping_dict(config['balance_mapping']['subjects'])
    INCOME_MAPPING = build_mapping_dict(config['income_mapping']['subjects'])
    REVENUE_MAPPING = build_mapping_dict(config['revenue_mapping']['subjects'])

    # 复合科目映射
    COMPOSITE_MAPPING = {}
    for comp_name, comp_config in config['composite_mapping']['subjects'].items():
        COMPOSITE_MAPPING[comp_name] = comp_config['components']

    print(f"✅ 成功加载科目映射配置文件: {CONFIG_FILE}")

except Exception as e:
    print(f"❌ 加载科目映射配置失败: {e}")
    print("使用默认配置...")

    # 回退到默认配置
    config = {"meta": {"version": "fallback", "description": "默认配置"}}
    BALANCE_MAPPING = {
        '资产': ['资产', '资产总计', '总资产'],
        '货币资金': ['货币资金', '现金及现金等价物', '现金'],
        '所有者权益': ['所有者权益', '股东权益', '权益合计', '权益总计'],
        '实收资本': ['实收资本', '实收资本（或股本）', '股本'],
    }

    INCOME_MAPPING = {
        '销售费用': ['销售费用'],
        '管理费用': ['管理费用'],
        '研发费用': ['研发费用'],
        '利息支出': ['利息支出'],
        '所得税费用': ['所得税费用', '所得税'],
    }

    REVENUE_MAPPING = {
        '主营业务收入': ['主营业务收入', '收入', '主营收入'],
        '主营业务成本': ['主营业务成本', '成本', '主营成本'],
    }

    COMPOSITE_MAPPING = {
        '折旧摊销': ['固定资产折旧', '使用权资产折旧', '无形资产摊销', '长期待摊费用摊销'],
        '有息负债': ['短期借款', '长期借款'],
    }


def find_row_by_label(sheet, label_mapping, target_label):
    """
    根据科目名称查找行号（支持多个别名匹配）

    Args:
        sheet: openpyxl工作表对象
        label_mapping: 科目映射字典
        target_label: 目标科目名称

    Returns:
        行号（1-based），如果未找到返回None
    """
    if target_label not in label_mapping:
        return None

    aliases = label_mapping[target_label]  # 获取所有别名

    # 遍历A列和B列查找匹配的科目（支持多个别名）
    for row in range(1, 200):  # 假设数据不超过200行
        # 检查A列
        cell_value_a = sheet.cell(row, 1).value
        if cell_value_a is not None:
            cell_value_a = str(cell_value_a).strip()
            for alias in aliases:
                if cell_value_a == alias or alias in cell_value_a:
                    return row

        # 检查B列（针对balance表）
        cell_value_b = sheet.cell(row, 2).value
        if cell_value_b is not None:
            cell_value_b = str(cell_value_b).strip()
            for alias in aliases:
                if cell_value_b == alias or alias in cell_value_b:
                    return row

    return None


def find_composite_rows(sheet, label_mapping, composite_label):
    """
    查找复合科目的所有行号

    Args:
        sheet: openpyxl工作表对象
        label_mapping: 科目映射字典
        composite_label: 复合科目名称（如'折旧摊销'）

    Returns:
        行号列表
    """
    if composite_label not in COMPOSITE_MAPPING:
        return []

    component_labels = COMPOSITE_MAPPING[composite_label]
    rows = []

    for component in component_labels:
        row = find_row_by_label(sheet, label_mapping, component)
        if row:
            rows.append(row)

    return rows


def get_value_by_label(sheet, label_mapping, target_label, col_idx):
    """
    根据科目名称获取数值

    Args:
        sheet: openpyxl工作表对象
        label_mapping: 科目映射字典
        target_label: 目标科目名称
        col_idx: 列索引

    Returns:
        数值，如果未找到返回0
    """
    # 检查是否是复合科目
    if target_label in COMPOSITE_MAPPING:
        rows = find_composite_rows(sheet, label_mapping, target_label)
        total = 0
        for row in rows:
            val = sheet.cell(row, col_idx).value
            if val is not None:
                total += val
        return total

    # 普通科目
    row = find_row_by_label(sheet, label_mapping, target_label)
    if row is None:
        return 0

    val = sheet.cell(row, col_idx).value
    return val if val is not None else 0


def print_mapping_report(sheet, label_mapping, sheet_name):
    """
    打印科目映射报告，用于调试

    Args:
        sheet: openpyxl工作表对象
        label_mapping: 科目映射字典
        sheet_name: 工作表名称
    """
    print(f"\n{'='*60}")
    print(f" {sheet_name} 科目映射报告")
    print('='*60)

    found_count = 0
    total_count = len(label_mapping)

    for target_label, aliases in label_mapping.items():
        row = find_row_by_label(sheet, label_mapping, target_label)
        if row:
            found_count += 1
            # 获取找到的实际科目名称
            actual_name_a = sheet.cell(row, 1).value
            actual_name_b = sheet.cell(row, 2).value
            actual_name = actual_name_a if actual_name_a else actual_name_b
            print(f"✓ {target_label:<20} -> 行{row:<3} ({actual_name})")
        else:
            print(f"✗ {target_label:<20} -> 未找到 (别名: {', '.join(aliases)})")

    print(f"\n统计: {found_count}/{total_count} 个科目找到 ({found_count/total_count*100:.1f}%)")


def get_config_info():
    """
    获取配置文件信息

    Returns:
        dict: 配置文件元信息
    """
    try:
        global config
        return config.get('meta', {})
    except:
        return {"version": "unknown", "description": "默认配置"}


def update_config(new_config):
    """
    更新配置文件

    Args:
        new_config: dict, 新的配置数据
    """
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(new_config, f, ensure_ascii=False, indent=2)
        print(f"✅ 配置文件已更新: {CONFIG_FILE}")

        # 重新加载配置到全局变量
        global config, BALANCE_MAPPING, INCOME_MAPPING, REVENUE_MAPPING, COMPOSITE_MAPPING
        config = load_config()
        BALANCE_MAPPING = build_mapping_dict(config['balance_mapping']['subjects'])
        INCOME_MAPPING = build_mapping_dict(config['income_mapping']['subjects'])
        REVENUE_MAPPING = build_mapping_dict(config['revenue_mapping']['subjects'])

        COMPOSITE_MAPPING = {}
        for comp_name, comp_config in config['composite_mapping']['subjects'].items():
            COMPOSITE_MAPPING[comp_name] = comp_config['components']

    except Exception as e:
        print(f"❌ 更新配置文件失败: {e}")


def add_subject_alias(mapping_type, subject_name, new_alias):
    """
    为指定科目添加新别名

    Args:
        mapping_type: str, 映射类型 ('balance', 'income', 'revenue')
        subject_name: str, 科目标准名称
        new_alias: str, 新别名
    """
    try:
        global config
        config_key = f"{mapping_type}_mapping"
        if config_key in config and subject_name in config[config_key]['subjects']:
            if new_alias not in config[config_key]['subjects'][subject_name]['aliases']:
                config[config_key]['subjects'][subject_name]['aliases'].append(new_alias)
                update_config(config)
                print(f"✅ 已为'{subject_name}'添加别名'{new_alias}'")
            else:
                print(f"⚠️ 别名'{new_alias}'已存在于'{subject_name}'")
        else:
            print(f"❌ 科目'{subject_name}'在{mapping_type}映射中不存在")
    except Exception as e:
        print(f"❌ 添加别名失败: {e}")


# 主程序：测试科目映射
if __name__ == '__main__':
    import sys
    import openpyxl

    if len(sys.argv) < 2:
        print("用法: python3 subject_mapping.py <Excel文件路径>")
        print("配置文件管理:")
        print("  --config-info       查看配置文件信息")
        print("  --add-alias <mapping_type> <subject> <alias>  添加别名")
        sys.exit(1)

    if sys.argv[1] == '--config-info':
        info = get_config_info()
        print(f"\n配置文件信息:")
        print(f"  版本: {info.get('version', 'N/A')}")
        print(f"  描述: {info.get('description', 'N/A')}")
        print(f"  更新时间: {info.get('last_updated', 'N/A')}")
        print(f"  文件路径: {CONFIG_FILE}")
        sys.exit(0)

    if sys.argv[1] == '--add-alias':
        if len(sys.argv) != 5:
            print("用法: python3 subject_mapping.py --add-alias <mapping_type> <subject> <alias>")
            print("示例: python3 subject_mapping.py --add-alias balance 货币资金 现金资产")
            sys.exit(1)
        mapping_type, subject, alias = sys.argv[2], sys.argv[3], sys.argv[4]
        add_subject_alias(mapping_type, subject, alias)
        sys.exit(0)

    input_file = sys.argv[1]

    try:
        print(f"正在分析文件: {input_file}")
        wb = openpyxl.load_workbook(input_file, data_only=True)

        if 'balance' in wb.sheetnames:
            print_mapping_report(wb['balance'], BALANCE_MAPPING, "资产负债表")

        if '损益现金流' in wb.sheetnames:
            print_mapping_report(wb['损益现金流'], INCOME_MAPPING, "损益现金流表")

        if '收入成本' in wb.sheetnames:
            print_mapping_report(wb['收入成本'], REVENUE_MAPPING, "收入成本表")

        wb.close()

    except Exception as e:
        print(f"❌ 分析过程出错: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)