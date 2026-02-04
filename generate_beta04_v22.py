#!/usr/bin/env python3
"""
Beta_04 财务分析表生成器 V2.2
基于动态科目识别，支持不同公司的数据结构

使用方法：
    python generate_beta04_v22.py input.xlsx output.xlsx
"""

import sys
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 导入科目映射和验证模块
try:
    from subject_mapping import get_value_by_label, BALANCE_MAPPING, INCOME_MAPPING, REVENUE_MAPPING, load_config
    from subject_validator_v22 import validate_subjects, print_validation_report
    MAPPING_AVAILABLE = True
    VALIDATION_AVAILABLE = True

    # 从JSON配置文件读取Beta_04字段映射
    config = load_config()
    BETA04_MAPPING = config.get('beta04_field_mapping', {})

except ImportError as e:
    print(f"警告: 模块导入失败 {e}")
    MAPPING_AVAILABLE = False
    VALIDATION_AVAILABLE = False

    # 回退到默认映射
    BETA04_MAPPING = {
        'balance': {},
        'income': {},
        'revenue': {}
    }

# 颜色定义
COLOR_INPUT = 'FF0000FF'       # 蓝色：手动输入
COLOR_FORMULA = 'FF000000'     # 黑色：公式计算
COLOR_LINK = 'FF008000'        # 绿色：工作表间引用
COLOR_HIGHLIGHT = 'FFFFFF00'   # 黄色背景：关键假设


def safe_divide(a, b, default=0):
    """安全除法"""
    try:
        if b is None or b == 0:
            return default
        return a / b
    except:
        return default


def safe_calc(func, default=0):
    """安全计算"""
    try:
        result = func()
        if result is None or str(result) == 'nan':
            return default
        return result
    except:
        return default


def get_dynamic_value(sheet, mapping_dict, field_name, col_idx):
    """
    动态获取科目数值（支持复合科目）

    Args:
        sheet: 工作表对象
        mapping_dict: 科目映射字典（BALANCE_MAPPING, INCOME_MAPPING, REVENUE_MAPPING）
        field_name: Beta_04字段名称
        col_idx: 列索引

    Returns:
        数值，如果找不到返回0
    """
    # 根据工作表名称确定映射类型
    sheet_name = sheet.title
    if sheet_name == 'balance':
        field_mapping = BETA04_MAPPING.get('balance', {})
    elif sheet_name == '损益现金流':
        field_mapping = BETA04_MAPPING.get('income', {})
    elif sheet_name == '收入成本':
        field_mapping = BETA04_MAPPING.get('revenue', {})
    else:
        return 0

    if field_name not in field_mapping:
        return 0

    subject_info = field_mapping[field_name]

    # 复合科目处理（支持list和tuple格式）
    if isinstance(subject_info, (list, tuple)):
        total = 0
        for subject in subject_info:
            value = get_value_by_label(sheet, mapping_dict, subject, col_idx)
            if value is not None:
                total += value
        return total

    # 单一科目处理
    return get_value_by_label(sheet, mapping_dict, subject_info, col_idx)


def create_beta04_analysis_v22(source_file, output_file):
    """创建 Beta_04 分析表（V2.2版本 - 动态科目识别）"""

    print(f"加载源文件: {source_file}")
    source_wb = load_workbook(source_file, data_only=True)

    # 检查必需的工作表
    required_sheets = ['balance', '损益现金流', '收入成本']
    for sheet_name in required_sheets:
        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"源文件缺少必需的工作表: {sheet_name}")

    # 验证科目完整性
    if VALIDATION_AVAILABLE:
        print("\n正在验证科目完整性...")
        is_valid, validation_report = validate_subjects(source_wb)

        if not is_valid:
            print_validation_report(validation_report)
            source_wb.close()
            raise ValueError("科目验证失败，请完善数据源后重试")

        print("✅ 科目验证通过：所有必需字段都已找到\n")

    # 获取源数据表
    balance = source_wb['balance']
    income = source_wb['损益现金流']
    revenue = source_wb['收入成本']

    # 检测年份列
    years = []
    balance_cols_idx = []  # balance表的列索引（从3开始）
    income_cols_idx = []   # 损益现金流表的列索引（从2开始）
    revenue_cols_idx = []  # 收入成本表的列索引（从2开始）

    for col_idx in range(3, 10):
        col_letter = get_column_letter(col_idx)
        cell_value = balance[f'{col_letter}1'].value
        if cell_value and '20' in str(cell_value):
            year = str(cell_value).split('.')[0]
            years.append(year)
            balance_cols_idx.append(col_idx)
            income_cols_idx.append(col_idx - 1)  # income表比balance表提前1列
            revenue_cols_idx.append(col_idx - 1)  # revenue表比balance表提前1列
            print(f"  发现年份: {year} (balance列{col_idx}, income/revenue列{col_idx-1})")

    if not years:
        raise ValueError("未找到年份数据")

    print(f"\n开始生成 Beta_04 分析表（共 {len(years)} 个年度）...")

    # 创建新工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = 'Beta_04'

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    for i in range(len(years)):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 15

    # 写入年份标题（第1行）
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # B1和C1设置为"定价"
    for col in ['B', 'C']:
        cell = ws[f'{col}1']
        cell.value = '定价'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

    # 年份标题
    for i, year in enumerate(years):
        col_letter = get_column_letter(4 + i)
        cell = ws[f'{col_letter}1']
        cell.value = year
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color='D0E0F0', end_color='D0E0F0', fill_type='solid')

    current_row = 2  # 从第2行开始写入数据
    row_map = {}  # 记录每个数据项的行号，用于公式引用

    def write_data_row(label, values, is_percentage=False, decimals=0,
                      is_input=False, color=None, formulas=None, is_bold=False):
        """写入数据行的通用函数"""
        nonlocal current_row

        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 写入标签
        label_cell = ws.cell(current_row, 1)
        label_cell.value = label
        label_cell.font = Font(bold=is_bold)
        label_cell.alignment = Alignment(horizontal='left', vertical='center')
        label_cell.border = thin_border

        # 写入数据或公式（从第4列开始，即D列）
        for i in range(len(years)):
            col_idx = 4 + i
            cell = ws.cell(current_row, col_idx)

            if formulas and i < len(formulas) and formulas[i]:
                # 写入公式
                cell.value = formulas[i]
                if color:
                    cell.font = Font(color=color)
            elif values and i < len(values):
                # 写入数值
                value = values[i]
                if is_percentage and value is not None:
                    cell.value = value
                    cell.number_format = '0.0%'
                else:
                    cell.value = value
                    if decimals == 0:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = f'#,##0.{"0" * decimals}'

                if is_input:
                    cell.font = Font(color=COLOR_INPUT)
                elif color:
                    cell.font = Font(color=color)
                else:
                    cell.font = Font(color=COLOR_LINK)

            # 设置对齐和边框
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = thin_border

        row_num = current_row
        current_row += 1
        return row_num

    # 数据提取阶段 - 使用动态科目识别
    print("正在提取数据...")
    data = {}

    for i, year in enumerate(years):
        data[year] = {}
        balance_col = balance_cols_idx[i]
        income_col = income_cols_idx[i]
        revenue_col = revenue_cols_idx[i]

        # 从资产负债表提取数据
        for field_name in BETA04_MAPPING['balance']:
            data[year][field_name] = get_dynamic_value(balance, BALANCE_MAPPING, field_name, balance_col)

        # 从损益现金流表提取数据
        for field_name in BETA04_MAPPING['income']:
            data[year][field_name] = get_dynamic_value(income, INCOME_MAPPING, field_name, income_col)

        # 从收入成本表提取数据
        for field_name in BETA04_MAPPING['revenue']:
            data[year][field_name] = get_dynamic_value(revenue, REVENUE_MAPPING, field_name, revenue_col)

        # 计算复合字段
        data[year]['费用'] = (data[year]['销售费用'] or 0) + (data[year]['管理费用'] or 0) + (data[year]['研发费用'] or 0) + (data[year]['税金及附加'] or 0)

    print("正在生成分析表...")

    # 第一部分：资产负债表数据（第2-26行）
    # 资产部分
    row_map['资产'] = write_data_row('资产', [data[year]['资产'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['现金'] = write_data_row('现金', [data[year]['现金'] for year in years], color=COLOR_LINK)
    row_map['投资'] = write_data_row('投资', [data[year]['投资'] for year in years], color=COLOR_LINK)
    row_map['运营资产'] = write_data_row('运营资产', [data[year]['运营资产'] for year in years], color=COLOR_LINK)
    row_map['应收'] = write_data_row('应收', [data[year]['应收'] for year in years], color=COLOR_LINK)
    row_map['预付'] = write_data_row('预付', [data[year]['预付'] for year in years], color=COLOR_LINK)
    row_map['存货'] = write_data_row('存货', [data[year]['存货'] for year in years], color=COLOR_LINK)
    row_map['长期资产'] = write_data_row('长期资产', [data[year]['长期资产'] for year in years], color=COLOR_LINK)

    # 负债部分
    row_map['负债'] = write_data_row('负债', [data[year]['负债'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['预收'] = write_data_row('预收', [data[year]['预收'] for year in years], color=COLOR_LINK)
    row_map['应付'] = write_data_row('应付', [data[year]['应付'] for year in years], color=COLOR_LINK)
    row_map['有息负债'] = write_data_row('有息负债', [data[year]['有息负债'] for year in years], color=COLOR_LINK)

    # 权益部分
    row_map['权益'] = write_data_row('权益', [data[year]['权益'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['资本投入'] = write_data_row('资本投入', [data[year]['资本投入'] for year in years], color=COLOR_LINK)
    row_map['未分配利润'] = write_data_row('未分配利润', [data[year]['未分配利润'] for year in years], color=COLOR_LINK)
    row_map['股息分红'] = write_data_row('股息分红', [data[year]['股息分红'] for year in years], color=COLOR_LINK, is_bold=True)

    current_row += 1  # 空行

    # 现金结构
    row_map['现金结构'] = write_data_row('现金结构', [data[year]['现金'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['利润留存_现金'] = write_data_row('利润留存', [data[year]['未分配利润'] for year in years], color=COLOR_LINK)
    # 股东、负债 = 现金 - 未分配利润
    股东负债_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        股东负债_formulas.append(f"={col}{row_map['现金']}-{col}{row_map['未分配利润']}")
    write_data_row('股东、负债', [None] * len(years), formulas=股东负债_formulas, color=COLOR_FORMULA)

    current_row += 1  # 空行

    # 折旧摊销和资本开支
    row_map['折旧摊销'] = write_data_row('折旧摊销', [data[year]['折旧摊销'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['资本开支'] = write_data_row('资本开支', [data[year]['资本开支'] for year in years], color=COLOR_LINK, is_bold=True)

    current_row += 1  # 空行

    # 股本
    row_map['股本'] = write_data_row('股本', [data[year]['股本'] for year in years], color=COLOR_LINK, is_bold=True)

    current_row += 1  # 空行

    # 第二部分：损益表数据（第28-46行）
    row_map['主营收入'] = write_data_row('主营收入', [data[year]['主营收入'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['主营成本'] = write_data_row('主营成本', [data[year]['主营成本'] for year in years], color=COLOR_LINK)
    row_map['费用'] = write_data_row('费用', [data[year]['费用'] for year in years], color=COLOR_LINK)
    row_map['研发费用'] = write_data_row('研发费用', [data[year]['研发费用'] for year in years], color=COLOR_LINK)
    row_map['销售费用'] = write_data_row('销售费用', [data[year]['销售费用'] for year in years], color=COLOR_LINK)
    row_map['管理费用'] = write_data_row('管理费用', [data[year]['管理费用'] for year in years], color=COLOR_LINK)
    row_map['税金及附加'] = write_data_row('税金及附加', [data[year]['税金及附加'] for year in years], color=COLOR_LINK)
    row_map['损耗'] = write_data_row('损耗', [data[year]['损耗'] for year in years], color=COLOR_LINK)
    row_map['其他收益'] = write_data_row('其他收益', [data[year]['其他收益'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['现金收益'] = write_data_row('现金收益', [data[year]['现金收益'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['股权收益'] = write_data_row('股权收益', [data[year]['股权收益'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['利息支出'] = write_data_row('利息支出', [data[year]['利息支出'] for year in years], color=COLOR_LINK, is_bold=True)
    row_map['财务费用'] = write_data_row('财务费用', [data[year]['财务费用'] for year in years], color=COLOR_LINK, is_bold=True)

    # 其他业务
    write_data_row('其他业务', [None] * len(years), is_bold=True)
    row_map['其他业务收入'] = write_data_row('收入', [data[year]['其他业务收入'] for year in years], color=COLOR_LINK)
    row_map['其他业务成本'] = write_data_row('成本', [data[year]['其他业务成本'] for year in years], color=COLOR_LINK)

    # 营业外
    write_data_row('营业外', [None] * len(years), is_bold=True)
    row_map['营业外收入'] = write_data_row('收入', [data[year]['营业外收入'] for year in years], color=COLOR_LINK)
    row_map['营业外成本'] = write_data_row('成本', [data[year]['营业外成本'] for year in years], color=COLOR_LINK)
    row_map['所得税'] = write_data_row('所得税', [data[year]['所得税'] for year in years], color=COLOR_LINK, is_bold=True)

    current_row += 2  # 两个空行

    # 第三部分：盈利结构分析（从第49行开始）
    write_data_row('盈利结构', [None] * len(years), is_bold=True)

    # 使用公式引用前面提取的数据
    主营收入_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        主营收入_盈利_formulas.append(f"={col}{row_map['主营收入']}")
    row_map['主营收入_盈利'] = write_data_row('主营收入', [None] * len(years),
                                           formulas=主营收入_盈利_formulas, color=COLOR_FORMULA, is_bold=True)

    主营成本_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        主营成本_盈利_formulas.append(f"={col}{row_map['主营成本']}")
    row_map['主营成本_盈利'] = write_data_row('主营成本', [None] * len(years),
                                           formulas=主营成本_盈利_formulas, color=COLOR_FORMULA)

    费用_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        费用_盈利_formulas.append(f"={col}{row_map['费用']}")
    row_map['费用_盈利'] = write_data_row('费用', [None] * len(years),
                                       formulas=费用_盈利_formulas, color=COLOR_FORMULA)

    损耗_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        损耗_盈利_formulas.append(f"={col}{row_map['损耗']}")
    row_map['损耗_盈利'] = write_data_row('损耗', [None] * len(years),
                                       formulas=损耗_盈利_formulas, color=COLOR_FORMULA)

    其他收益_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        其他收益_盈利_formulas.append(f"={col}{row_map['其他收益']}")
    row_map['其他收益_盈利'] = write_data_row('其他收益', [None] * len(years),
                                           formulas=其他收益_盈利_formulas, color=COLOR_FORMULA)

    现金收益_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        现金收益_盈利_formulas.append(f"={col}{row_map['现金收益']}")
    row_map['现金收益_盈利'] = write_data_row('现金收益', [None] * len(years),
                                           formulas=现金收益_盈利_formulas, color=COLOR_FORMULA)

    利息支出_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        利息支出_盈利_formulas.append(f"={col}{row_map['利息支出']}")
    row_map['利息支出_盈利'] = write_data_row('利息支出', [None] * len(years),
                                           formulas=利息支出_盈利_formulas, color=COLOR_FORMULA)

    财务费用_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        财务费用_盈利_formulas.append(f"={col}{row_map['财务费用']}")
    row_map['财务费用_盈利'] = write_data_row('财务费用', [None] * len(years),
                                           formulas=财务费用_盈利_formulas, color=COLOR_FORMULA)

    所得税_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        所得税_盈利_formulas.append(f"={col}{row_map['所得税']}")
    row_map['所得税_盈利'] = write_data_row('所得税', [None] * len(years),
                                         formulas=所得税_盈利_formulas, color=COLOR_FORMULA)

    # 业务利润公式
    业务利润_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        formula = (f"={col}{row_map['主营收入_盈利']}-{col}{row_map['主营成本_盈利']}-{col}{row_map['费用_盈利']}"
                  f"+{col}{row_map['损耗_盈利']}+{col}{row_map['其他收益_盈利']}+{col}{row_map['现金收益_盈利']}"
                  f"-{col}{row_map['利息支出_盈利']}-{col}{row_map['财务费用_盈利']}-{col}{row_map['所得税_盈利']}")
        业务利润_formulas.append(formula)

    row_map['业务利润'] = write_data_row('业务利润', [None] * len(years),
                                       formulas=业务利润_formulas, color=COLOR_FORMULA, is_bold=True)

    # 股权收益
    股权收益_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        股权收益_盈利_formulas.append(f"={col}{row_map['股权收益']}")
    row_map['股权收益_盈利'] = write_data_row('股权收益', [None] * len(years),
                                           formulas=股权收益_盈利_formulas, color=COLOR_FORMULA)

    # 其他业务利润公式
    其他业务利润_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        formula = f"={col}{row_map['其他业务收入']}-{col}{row_map['其他业务成本']}"
        其他业务利润_formulas.append(formula)
    row_map['其他业务利润'] = write_data_row('其他业务利润', [None] * len(years),
                                         formulas=其他业务利润_formulas, color=COLOR_FORMULA)

    # 营业外利润公式
    营业外利润_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        formula = f"={col}{row_map['营业外收入']}-{col}{row_map['营业外成本']}"
        营业外利润_formulas.append(formula)
    row_map['营业外利润'] = write_data_row('营业外利润', [None] * len(years),
                                        formulas=营业外利润_formulas, color=COLOR_FORMULA)

    # 净利润公式
    净利润_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        formula = f"={col}{row_map['业务利润']}+{col}{row_map['股权收益_盈利']}+{col}{row_map['其他业务利润']}+{col}{row_map['营业外利润']}"
        净利润_formulas.append(formula)
    row_map['净利润'] = write_data_row('净利润', [None] * len(years),
                                    formulas=净利润_formulas, color=COLOR_FORMULA)

    # 分红
    分红_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        分红_formulas.append(f"={col}{row_map['股息分红']}")
    row_map['分红'] = write_data_row('分红', [None] * len(years),
                                   formulas=分红_formulas, color=COLOR_FORMULA)

    # 利润留存公式
    利润留存_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        formula = f"={col}{row_map['净利润']}-{col}{row_map['分红']}"
        利润留存_formulas.append(formula)
    row_map['利润留存'] = write_data_row('利润留存', [None] * len(years),
                                      formulas=利润留存_formulas, color=COLOR_FORMULA, is_bold=True)

    # 折旧摊销
    折旧摊销_盈利_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        折旧摊销_盈利_formulas.append(f"={col}{row_map['折旧摊销']}")
    row_map['折旧摊销_盈利'] = write_data_row('折旧摊销', [None] * len(years),
                                           formulas=折旧摊销_盈利_formulas, color=COLOR_FORMULA)

    # CAPEX
    CAPEX_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        CAPEX_formulas.append(f"={col}{row_map['资本开支']}")
    row_map['CAPEX'] = write_data_row('CAPEX', [None] * len(years),
                                     formulas=CAPEX_formulas, color=COLOR_FORMULA)

    # 现金留存公式
    现金留存_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        formula = f"={col}{row_map['利润留存']}+{col}{row_map['折旧摊销_盈利']}-{col}{row_map['CAPEX']}"
        现金留存_formulas.append(formula)
    row_map['现金留存'] = write_data_row('现金留存', [None] * len(years),
                                      formulas=现金留存_formulas, color=COLOR_FORMULA, is_bold=True)

    # 业务现金公式（业务利润 + 折旧摊销 - 资本开支）
    业务现金_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        formula = f"={col}{row_map['业务利润']}+{col}{row_map['折旧摊销_盈利']}-{col}{row_map['CAPEX']}"
        业务现金_formulas.append(formula)
    row_map['业务现金'] = write_data_row('业务现金', [None] * len(years),
                                      formulas=业务现金_formulas, color=COLOR_FORMULA, is_bold=True)

    current_row += 1  # 空行

    # 第四部分：经营指标（从第71行开始）
    write_data_row('经营周转', [None] * len(years), is_bold=True)

    # 毛利率
    毛利率_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        毛利率_formulas.append(f"=({col}{row_map['主营收入']}-{col}{row_map['主营成本']})/{col}{row_map['主营收入']}")
    write_data_row('毛利率', [None] * len(years), is_percentage=True,
                  formulas=毛利率_formulas, color=COLOR_FORMULA, is_bold=True)

    # 费率
    费率_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        费率_formulas.append(f"={col}{row_map['费用']}/{col}{row_map['主营收入']}")
    write_data_row('费率', [None] * len(years), is_percentage=True,
                  formulas=费率_formulas, color=COLOR_FORMULA, is_bold=True)

    # 研发费率
    研发费率_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        研发费率_formulas.append(f"={col}{row_map['研发费用']}/{col}{row_map['主营收入']}")
    write_data_row('研发费率', [None] * len(years), is_percentage=True,
                  formulas=研发费率_formulas, color=COLOR_FORMULA, is_bold=True)

    # 销售费率
    销售费率_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        销售费率_formulas.append(f"={col}{row_map['销售费用']}/{col}{row_map['主营收入']}")
    write_data_row('销售费率', [None] * len(years), is_percentage=True,
                  formulas=销售费率_formulas, color=COLOR_FORMULA, is_bold=True)

    # 管理费率
    管理费率_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        管理费率_formulas.append(f"={col}{row_map['管理费用']}/{col}{row_map['主营收入']}")
    write_data_row('管理费率', [None] * len(years), is_percentage=True,
                  formulas=管理费率_formulas, color=COLOR_FORMULA, is_bold=True)

    # 净利率
    净利率_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        净利率_formulas.append(f"={col}{row_map['业务利润']}/{col}{row_map['主营收入']}")
    write_data_row('净利率', [None] * len(years), is_percentage=True,
                  formulas=净利率_formulas, color=COLOR_FORMULA, is_bold=True)

    current_row += 1  # 空行

    # 应收占比
    应收占比_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        应收占比_formulas.append(f"={col}{row_map['应收']}/{col}{row_map['主营收入']}")
    write_data_row('应收占比', [None] * len(years), is_percentage=True,
                  formulas=应收占比_formulas, color=COLOR_FORMULA, is_bold=True)

    # 应收周转
    应收周转_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        应收周转_formulas.append(f"={col}{row_map['主营收入']}/{col}{row_map['应收']}")
    write_data_row('应收周转', [None] * len(years), decimals=1,
                  formulas=应收周转_formulas, color=COLOR_FORMULA, is_bold=True)

    # 存货周转
    存货周转_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        存货周转_formulas.append(f"={col}{row_map['主营成本']}/{col}{row_map['存货']}")
    write_data_row('存货周转', [None] * len(years), decimals=1,
                  formulas=存货周转_formulas, color=COLOR_FORMULA, is_bold=True)

    # 应付（预付）占比
    应付预付占比_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        应付预付占比_formulas.append(f"=({col}{row_map['应付']}+{col}{row_map['预付']})/{col}{row_map['主营收入']}")
    write_data_row('应付（预付）占比', [None] * len(years), is_percentage=True,
                  formulas=应付预付占比_formulas, color=COLOR_FORMULA, is_bold=True)

    # 预收占比
    预收占比_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        预收占比_formulas.append(f"={col}{row_map['预收']}/{col}{row_map['主营收入']}")
    write_data_row('预收占比', [None] * len(years), is_percentage=True,
                  formulas=预收占比_formulas, color=COLOR_FORMULA, is_bold=True)

    # 利息支出/业务利润
    利息业务利润比_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        利息业务利润比_formulas.append(f"={col}{row_map['利息支出']}/{col}{row_map['业务利润']}")
    write_data_row('利息支出/业务利润', [None] * len(years), is_percentage=True,
                  formulas=利息业务利润比_formulas, color=COLOR_FORMULA, is_bold=True)

    # 折旧摊销/长期资产
    折旧长期资产比_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        折旧长期资产比_formulas.append(f"={col}{row_map['折旧摊销']}/{col}{row_map['长期资产']}")
    write_data_row('折旧摊销/长期资产', [None] * len(years), is_percentage=True,
                  formulas=折旧长期资产比_formulas, color=COLOR_FORMULA, is_bold=True)

    current_row += 1  # 空行

    # 第五部分：资产负债结构（从第87行开始）
    write_data_row('资产结构', [None] * len(years), is_bold=True)

    # 现金资产
    现金资产_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        现金资产_formulas.append(f"={col}{row_map['现金']}/{col}{row_map['资产']}")
    write_data_row('现金资产', [None] * len(years), is_percentage=True,
                  formulas=现金资产_formulas, color=COLOR_FORMULA, is_bold=True)

    # 股权资产
    股权资产_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        股权资产_formulas.append(f"={col}{row_map['投资']}/{col}{row_map['资产']}")
    write_data_row('股权资产', [None] * len(years), is_percentage=True,
                  formulas=股权资产_formulas, color=COLOR_FORMULA, is_bold=True)

    # 经营资产
    经营资产_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        经营资产_formulas.append(f"={col}{row_map['运营资产']}/{col}{row_map['资产']}")
    write_data_row('经营资产', [None] * len(years), is_percentage=True,
                  formulas=经营资产_formulas, color=COLOR_FORMULA, is_bold=True)

    # 长期资产
    长期资产_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        长期资产_formulas.append(f"={col}{row_map['长期资产']}/{col}{row_map['资产']}")
    write_data_row('长期资产', [None] * len(years), is_percentage=True,
                  formulas=长期资产_formulas, color=COLOR_FORMULA, is_bold=True)

    # 负债结构
    write_data_row('负债结构', [None] * len(years), is_bold=True)

    # 经营负债
    经营负债_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        经营负债_formulas.append(f"=({col}{row_map['预收']}+{col}{row_map['应付']})/{col}{row_map['负债']}")
    write_data_row('经营负债', [None] * len(years), is_percentage=True,
                  formulas=经营负债_formulas, color=COLOR_FORMULA, is_bold=True)

    # 有息负债
    有息负债_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        有息负债_formulas.append(f"={col}{row_map['有息负债']}/{col}{row_map['负债']}")
    write_data_row('有息负债', [None] * len(years), is_percentage=True,
                  formulas=有息负债_formulas, color=COLOR_FORMULA, is_bold=True)

    # 权益结构
    write_data_row('权益结构', [None] * len(years), is_bold=True)

    # 资本投入
    资本投入_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        资本投入_formulas.append(f"={col}{row_map['资本投入']}/{col}{row_map['权益']}")
    write_data_row('资本投入', [None] * len(years), is_percentage=True,
                  formulas=资本投入_formulas, color=COLOR_FORMULA, is_bold=True)

    # 利润留存
    利润留存_权益_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        利润留存_权益_formulas.append(f"={col}{row_map['未分配利润']}/{col}{row_map['权益']}")
    write_data_row('利润留存', [None] * len(years), is_percentage=True,
                  formulas=利润留存_权益_formulas, color=COLOR_FORMULA, is_bold=True)

    current_row += 1  # 空行

    # 第六部分：同比增长分析（从第99行开始）

    # 营收YoY
    营收YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:  # 最早年份没有YoY
            营收YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            营收YoY_formulas.append(f"=({this_col}{row_map['主营收入']}-{prev_col}{row_map['主营收入']})/{prev_col}{row_map['主营收入']}")
    write_data_row('营收YoY', [None] * len(years), is_percentage=True,
                  formulas=营收YoY_formulas, color=COLOR_FORMULA, is_bold=True)

    # 业务净利润YoY
    业务净利润YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:  # 最早年份没有YoY
            业务净利润YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            业务净利润YoY_formulas.append(f"=({this_col}{row_map['业务利润']}-{prev_col}{row_map['业务利润']})/{prev_col}{row_map['业务利润']}")
    write_data_row('业务净利润YoY', [None] * len(years), is_percentage=True,
                  formulas=业务净利润YoY_formulas, color=COLOR_FORMULA, is_bold=True)

    current_row += 1  # 空行

    # 资产YoY
    资产YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:  # 最早年份没有YoY
            资产YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            资产YoY_formulas.append(f"=({this_col}{row_map['资产']}-{prev_col}{row_map['资产']})/{prev_col}{row_map['资产']}")
    write_data_row('资产YoY', [None] * len(years), is_percentage=True,
                  formulas=资产YoY_formulas, color=COLOR_FORMULA, is_bold=True)

    # 现金YoY
    现金YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            现金YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            现金YoY_formulas.append(f"=({this_col}{row_map['现金']}-{prev_col}{row_map['现金']})/{prev_col}{row_map['现金']}")
    write_data_row('现金', [None] * len(years), is_percentage=True,
                  formulas=现金YoY_formulas, color=COLOR_FORMULA)

    # 投资YoY
    投资YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            投资YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            投资YoY_formulas.append(f"=({this_col}{row_map['投资']}-{prev_col}{row_map['投资']})/{prev_col}{row_map['投资']}")
    write_data_row('投资', [None] * len(years), is_percentage=True,
                  formulas=投资YoY_formulas, color=COLOR_FORMULA)

    # 运营资产YoY
    运营资产YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            运营资产YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            运营资产YoY_formulas.append(f"=({this_col}{row_map['运营资产']}-{prev_col}{row_map['运营资产']})/{prev_col}{row_map['运营资产']}")
    write_data_row('运营资产', [None] * len(years), is_percentage=True,
                  formulas=运营资产YoY_formulas, color=COLOR_FORMULA)

    # 长期资产YoY
    长期资产YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            长期资产YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            长期资产YoY_formulas.append(f"=({this_col}{row_map['长期资产']}-{prev_col}{row_map['长期资产']})/{prev_col}{row_map['长期资产']}")
    write_data_row('长期资产', [None] * len(years), is_percentage=True,
                  formulas=长期资产YoY_formulas, color=COLOR_FORMULA)

    # 权益YoY
    权益YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            权益YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            权益YoY_formulas.append(f"=({this_col}{row_map['权益']}-{prev_col}{row_map['权益']})/{prev_col}{row_map['权益']}")
    write_data_row('权益YoY', [None] * len(years), is_percentage=True,
                  formulas=权益YoY_formulas, color=COLOR_FORMULA, is_bold=True)

    # 资本投入YoY
    资本投入YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            资本投入YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            资本投入YoY_formulas.append(f"=({this_col}{row_map['资本投入']}-{prev_col}{row_map['资本投入']})/{prev_col}{row_map['资本投入']}")
    write_data_row('资本投入', [None] * len(years), is_percentage=True,
                  formulas=资本投入YoY_formulas, color=COLOR_FORMULA)

    # 未分配利润YoY
    未分配利润YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            未分配利润YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            未分配利润YoY_formulas.append(f"=({this_col}{row_map['未分配利润']}-{prev_col}{row_map['未分配利润']})/{prev_col}{row_map['未分配利润']}")
    write_data_row('未分配利润', [None] * len(years), is_percentage=True,
                  formulas=未分配利润YoY_formulas, color=COLOR_FORMULA)

    # 负债YoY
    负债YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            负债YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            负债YoY_formulas.append(f"=({this_col}{row_map['负债']}-{prev_col}{row_map['负债']})/{prev_col}{row_map['负债']}")
    write_data_row('负债YoY', [None] * len(years), is_percentage=True,
                  formulas=负债YoY_formulas, color=COLOR_FORMULA, is_bold=True)

    # 预收YoY
    预收YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            预收YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            预收YoY_formulas.append(f"=({this_col}{row_map['预收']}-{prev_col}{row_map['预收']})/{prev_col}{row_map['预收']}")
    write_data_row('预收', [None] * len(years), is_percentage=True,
                  formulas=预收YoY_formulas, color=COLOR_FORMULA)

    # 应付YoY
    应付YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            应付YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            应付YoY_formulas.append(f"=({this_col}{row_map['应付']}-{prev_col}{row_map['应付']})/{prev_col}{row_map['应付']}")
    write_data_row('应付', [None] * len(years), is_percentage=True,
                  formulas=应付YoY_formulas, color=COLOR_FORMULA)

    # 有息债务YoY
    有息债务YoY_formulas = []
    for i in range(len(years)):
        if i == len(years) - 1:
            有息债务YoY_formulas.append(None)
        else:
            this_col = get_column_letter(4 + i)
            prev_col = get_column_letter(4 + i + 1)
            有息债务YoY_formulas.append(f"=({this_col}{row_map['有息负债']}-{prev_col}{row_map['有息负债']})/{prev_col}{row_map['有息负债']}")
    write_data_row('有息债务', [None] * len(years), is_percentage=True,
                  formulas=有息债务YoY_formulas, color=COLOR_FORMULA)

    current_row += 1  # 空行

    # 第七部分：ROE分解（从第115行开始）

    # ROE
    ROE_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        ROE_formulas.append(f"={col}{row_map['业务利润']}/{col}{row_map['权益']}")
    write_data_row('ROE', [None] * len(years), is_percentage=True,
                  formulas=ROE_formulas, color=COLOR_FORMULA, is_bold=True)

    # Net Profit（净利率）
    净利率_ROE_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        净利率_ROE_formulas.append(f"={col}{row_map['业务利润']}/{col}{row_map['主营收入']}")
    write_data_row('Net Proit', [None] * len(years), is_percentage=True,
                  formulas=净利率_ROE_formulas, color=COLOR_FORMULA)

    # 主营收入/资产（资产周转率）
    资产周转率_ROE_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        资产周转率_ROE_formulas.append(f"={col}{row_map['主营收入']}/{col}{row_map['资产']}")
    write_data_row('主营收入/资产', [None] * len(years), decimals=2,
                  formulas=资产周转率_ROE_formulas, color=COLOR_FORMULA)

    # 资产/资本（权益乘数）
    权益乘数_ROE_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        权益乘数_ROE_formulas.append(f"={col}{row_map['资产']}/{col}{row_map['权益']}")
    write_data_row('资产/资本', [None] * len(years), decimals=2,
                  formulas=权益乘数_ROE_formulas, color=COLOR_FORMULA)

    current_row += 1  # 空行

    # 第八部分：估值指标（从第120行开始）
    # EPS（每股收益）
    EPS_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        EPS_formulas.append(f"={col}{row_map['业务利润']}/{col}{row_map['股本']}")
    write_data_row('EPS', [None] * len(years), decimals=2,
                  formulas=EPS_formulas, color=COLOR_FORMULA, is_bold=True)

    # FPS（每股自由现金流）
    FPS_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        FPS_formulas.append(f"={col}{row_map['业务现金']}/{col}{row_map['股本']}")
    write_data_row('FPS', [None] * len(years), decimals=2,
                  formulas=FPS_formulas, color=COLOR_FORMULA, is_bold=True)

    current_row += 1  # 空行

    # 手动输入股价（年度最高价和最低价）
    row_map['价格H'] = write_data_row('价格（H）', [None] * len(years), decimals=2, is_input=True, color=COLOR_INPUT, is_bold=True)
    row_map['价格L'] = write_data_row('价格（L）', [None] * len(years), decimals=2, is_input=True, color=COLOR_INPUT, is_bold=True)

    # PE（高）
    PE_H_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        PE_H_formulas.append(f"={col}{row_map['价格H']}/({col}{row_map['业务利润']}/{col}{row_map['股本']})")
    write_data_row('PE', [None] * len(years), decimals=2,
                  formulas=PE_H_formulas, color=COLOR_FORMULA, is_bold=True)

    # PE（低）
    PE_L_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        PE_L_formulas.append(f"={col}{row_map['价格L']}/({col}{row_map['业务利润']}/{col}{row_map['股本']})")
    write_data_row('PE', [None] * len(years), decimals=2,
                  formulas=PE_L_formulas, color=COLOR_FORMULA, is_bold=True)

    # 股息率（高）- 分子取股息分红
    股息率_H_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        股息率_H_formulas.append(f"={col}{row_map['股息分红']}/({col}{row_map['价格H']}*{col}{row_map['股本']})")
    write_data_row('股息率', [None] * len(years), is_percentage=True,
                  formulas=股息率_H_formulas, color=COLOR_FORMULA, is_bold=True)

    # 股息率（低）- 分子取股息分红
    股息率_L_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        股息率_L_formulas.append(f"={col}{row_map['股息分红']}/({col}{row_map['价格L']}*{col}{row_map['股本']})")
    write_data_row('股息率', [None] * len(years), is_percentage=True,
                  formulas=股息率_L_formulas, color=COLOR_FORMULA, is_bold=True)

    current_row += 2  # 两个空行

    # 市值（高）
    市值_H_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        市值_H_formulas.append(f"={col}{row_map['价格H']}*{col}{row_map['股本']}")
    write_data_row('市值', [None] * len(years), decimals=0,
                  formulas=市值_H_formulas, color=COLOR_FORMULA, is_bold=True)

    # 市值（低）
    市值_L_formulas = []
    for i in range(len(years)):
        col = get_column_letter(4 + i)
        市值_L_formulas.append(f"={col}{row_map['价格L']}*{col}{row_map['股本']}")
    write_data_row('市值', [None] * len(years), decimals=0,
                  formulas=市值_L_formulas, color=COLOR_FORMULA, is_bold=True)

    # 保存文件
    print(f"\n保存文件: {output_file}")
    wb.save(output_file)
    source_wb.close()

    print("✓ Beta_04 分析表创建完成")

    # 显示关键指标
    if years:
        latest_year = years[0]
        营收 = data[latest_year]['主营收入']
        业务利润 = safe_calc(lambda: (data[latest_year]['主营收入'] or 0) -
                             (data[latest_year]['主营成本'] or 0) -
                             (data[latest_year]['费用'] or 0) +
                             (data[latest_year]['损耗'] or 0) +
                             (data[latest_year]['其他收益'] or 0) +
                             (data[latest_year]['现金收益'] or 0) -
                             (data[latest_year]['利息支出'] or 0) -
                             (data[latest_year]['财务费用'] or 0) -
                             (data[latest_year]['所得税'] or 0))
        毛利率 = safe_divide((营收 or 0) - (data[latest_year]['主营成本'] or 0), 营收 or 1) * 100
        净利率 = safe_divide(业务利润, 营收 or 1) * 100
        ROE = safe_divide(业务利润, data[latest_year]['权益'] or 1) * 100
        EPS = safe_divide(业务利润, data[latest_year]['股本'] or 1)

        print("=" * 60)
        print("✓ 成功生成 Beta_04 财务分析表")
        print(f"  输入文件: {source_file}")
        print(f"  输出文件: {output_file}")
        print(f"  分析年度: {', '.join(years)}")
        print("=" * 60)
        print(f"【{latest_year}年关键指标】")
        print(f"  营收: {int(营收):,}" if 营收 else "  营收: N/A")
        print(f"  业务利润: {int(业务利润):,}" if 业务利润 else "  业务利润: N/A")
        print(f"  毛利率: {毛利率:.1f}%")
        print(f"  净利率: {净利率:.1f}%")
        print(f"  ROE: {ROE:.1f}%")
        print(f"  EPS: {EPS:.2f}" if EPS else "  EPS: N/A")


def main():
    if len(sys.argv) < 2:
        print("用法: python generate_beta04_v22.py <输入文件.xlsx> [输出文件.xlsx]")
        print("示例: python generate_beta04_v22.py 福耀玻璃.xlsx 福耀玻璃_Beta04.xlsx")
        sys.exit(1)

    source = sys.argv[1]
    output = sys.argv[2] if len(sys.argv) > 2 else source.replace('.xlsx', '_Beta04_v22.xlsx')

    if not MAPPING_AVAILABLE:
        print("❌ 错误: 科目映射模块不可用，请检查 subject_mapping.py 文件")
        sys.exit(1)

    try:
        create_beta04_analysis_v22(source, output)
    except Exception as e:
        print(f"\n✗ 错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()