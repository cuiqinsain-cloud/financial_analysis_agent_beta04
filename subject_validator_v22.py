#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
科目完整性验证模块 - V2.2版本
基于动态科目识别的验证，覆盖所有beta04_field_mapping字段（仅严格模式）
"""

from subject_mapping import find_row_by_label, BALANCE_MAPPING, INCOME_MAPPING, REVENUE_MAPPING, load_config


def validate_subjects(wb):
    """
    全面验证科目是否能够被动态识别（V2.2版本 - 严格模式）

    基于beta04_field_mapping中所有字段进行验证

    参数:
        wb: openpyxl workbook 对象

    返回:
        (is_valid, validation_report) 元组
        - is_valid: bool, 是否所有必需字段都能找到
        - validation_report: dict, 详细验证报告
    """
    config = load_config()
    validation_rules = config.get('validation_rules', {})

    # 获取所有beta04_field_mapping中的字段（严格模式）
    balance_fields = validation_rules.get('balance_fields', {}).get('required', [])
    income_fields = validation_rules.get('income_fields', {}).get('required', [])
    revenue_fields = validation_rules.get('revenue_fields', {}).get('required', [])

    validation_report = {
        'balance': {
            'missing': [],
            'found': [],
            'total_required': len(balance_fields)
        },
        '损益现金流': {
            'missing': [],
            'found': [],
            'total_required': len(income_fields)
        },
        '收入成本': {
            'missing': [],
            'found': [],
            'total_required': len(revenue_fields)
        },
        'summary': {
            'total_missing': 0,
            'total_found': 0,
            'total_required': len(balance_fields) + len(income_fields) + len(revenue_fields)
        }
    }

    # 验证资产负债表字段
    if 'balance' in wb.sheetnames:
        balance_ws = wb['balance']
        for field in balance_fields:
            row = find_row_by_label(balance_ws, BALANCE_MAPPING, field)
            if row is None:
                validation_report['balance']['missing'].append(field)
            else:
                validation_report['balance']['found'].append(field)
    else:
        validation_report['balance']['missing'] = balance_fields

    # 验证损益现金流表字段
    if '损益现金流' in wb.sheetnames:
        income_ws = wb['损益现金流']
        for field in income_fields:
            row = find_row_by_label(income_ws, INCOME_MAPPING, field)
            if row is None:
                validation_report['损益现金流']['missing'].append(field)
            else:
                validation_report['损益现金流']['found'].append(field)
    else:
        validation_report['损益现金流']['missing'] = income_fields

    # 验证收入成本表字段
    if '收入成本' in wb.sheetnames:
        revenue_ws = wb['收入成本']
        for field in revenue_fields:
            row = find_row_by_label(revenue_ws, REVENUE_MAPPING, field)
            if row is None:
                validation_report['收入成本']['missing'].append(field)
            else:
                validation_report['收入成本']['found'].append(field)
    else:
        validation_report['收入成本']['missing'] = revenue_fields

    # 计算汇总统计
    total_missing = (
        len(validation_report['balance']['missing']) +
        len(validation_report['损益现金流']['missing']) +
        len(validation_report['收入成本']['missing'])
    )

    total_found = (
        len(validation_report['balance']['found']) +
        len(validation_report['损益现金流']['found']) +
        len(validation_report['收入成本']['found'])
    )

    validation_report['summary']['total_missing'] = total_missing
    validation_report['summary']['total_found'] = total_found

    # 判断是否验证通过
    is_valid = total_missing == 0

    return is_valid, validation_report


def print_validation_report(validation_report):
    """
    打印验证报告（V2.2版本 - 严格模式）
    """
    summary = validation_report['summary']

    print(f"\n📊 科目验证报告 (严格模式 - 全字段验证)")
    print("=" * 60)

    # 打印汇总统计
    completion_rate = (summary['total_found'] / summary['total_required'] * 100) if summary['total_required'] > 0 else 0
    print(f"总计: {summary['total_found']}/{summary['total_required']} 个字段找到 ({completion_rate:.1f}%)")
    print(f"缺失: {summary['total_missing']} 个字段")

    if summary['total_missing'] == 0:
        print("✅ 验证通过：所有必需字段都已找到\n")
        return

    print(f"❌ 验证失败：发现 {summary['total_missing']} 个缺失字段\n")

    # 详细报告每个工作表
    for sheet_name in ['balance', '损益现金流', '收入成本']:
        sheet_report = validation_report[sheet_name]
        cn_name = {'balance': '资产负债表', '损益现金流': '损益现金流表', '收入成本': '收入成本表'}[sheet_name]

        if sheet_report['missing']:
            completion_rate = ((sheet_report['total_required'] - len(sheet_report['missing'])) / sheet_report['total_required'] * 100) if sheet_report['total_required'] > 0 else 0
            print(f"【{cn_name}缺失字段 - {len(sheet_report['missing'])}/{sheet_report['total_required']} ({completion_rate:.1f}%完成度)】")

            for field in sheet_report['missing']:
                print(f"  ❌ {field}")
            print()

        if sheet_report['found']:
            print(f"【{cn_name}已找到字段 - {len(sheet_report['found'])}/{sheet_report['total_required']}】")

            for field in sheet_report['found']:
                print(f"  ✅ {field}")
            print()

    # 修复建议
    print("🔧 修复建议：")
    print("1. 确认源数据文件包含上述缺失字段")
    print("2. 检查字段名称是否能够被动态识别")
    print("3. 运行 'python3 subject_mapping.py 你的文件.xlsx' 查看详细映射报告")
    print("4. 考虑在 subject_mapping_config.json 中添加字段别名")
    print()


if __name__ == '__main__':
    import sys
    import openpyxl

    if len(sys.argv) < 2:
        print("用法: python3 subject_validator_v22.py <Excel文件路径>")
        print("说明: V2.2版本使用严格模式，验证所有beta04_field_mapping字段")
        sys.exit(1)

    input_file = sys.argv[1]

    try:
        print(f"正在验证文件 (V2.2 - 严格模式): {input_file}")
        wb = openpyxl.load_workbook(input_file, data_only=True)

        is_valid, validation_report = validate_subjects(wb)
        print_validation_report(validation_report)

        wb.close()

        if not is_valid:
            sys.exit(1)

    except Exception as e:
        print(f"❌ 验证过程出错: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)